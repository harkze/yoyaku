"""
2024 kujirahand
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the “Software”), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED “AS IS”, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

"""
import datetime
import os
import subprocess
import unicodedata

import openpyxl
import TkEasyGUI as sg


class Sercher():
    def __init__(self):
        self.excel_name = "./book/Book1.xlsx"
        self.wb = openpyxl.load_workbook(self.excel_name)
        self.ws = self.wb.active
        self.colum_name_list = []
        self.table = []
        self.row = 0
        self.int_to_alpha = {
            1:"A",
            2:"B",
            3:"C",
            4:"D",
            5:"E",
            6:"F",
            7:"G",
            8:"H",
            9:"I",
            10:"J",
            11:"K",
            12:"L",
            13:"M",
            14:"N",
            15:"O",
            16:"P",
            17:"Q",
            18:"R",
            19:"S",
            20:"T",
            21:"U"
        }

    def open(self):
        try:
            subprocess.Popen(['start',self.excel_name], shell=True)
        except:
            sg.print("Excelを閉じてください")

    def data_mining(self, colum_name, keyword):
        col = self.find_colum(colum_name, keyword)
        if not col:
            return []
        dic = {
            "No.":col[2],
            "名前":col[3],
            "注文方法":col[0],
            "備考":col[19],
            "会計未済・※":col[1],
            "cake":{}
        }
        if not col[2]:
            dic["No."] = "無記入"
        else:
            dic["No."] = col[2]
        if not col[3]:
            dic["名前"] = "無記入"
        else:
            dic["名前"] = col[3]
        if not col[0]:
            dic["注文方法"] = "無記入"
        else:
            dic["注文方法"] = col[0]
        if not col[19]:
            dic["備考"] = "なし"
        else:
            dic["備考"] = col[19]
        if not col[1]:
            dic["会計未済・※"] = "無記入"
        else:
            dic["会計未済・※"] = col[1]
        cake_dir = {
            6:"short4",
            7:"short5",
            8:"short6",
            9:"short7",
            10:"choco4",
            11:"choco5",
            12:"choco6",
            13:"choco7",
            14:"スノーモンブラン",
            15:"リッチショコラ",
            16:"紅タル",
            17:"フロマージュ"
        }
        tmp = []
        for i in range(6, 18):
            if col[i]:
                if int(col[i]) >= 1:
                    dic["cake"][cake_dir[i]] = col[i]
        return dic
    
    def find_cell_from_array(self, arr, want_colum_name):
        idx = self.colum_name_list.index(want_colum_name)
        return arr[idx]

    def find_cell_from_colum(self, keyword_colum_name, keyword, want_colum_name):
        idx = self.colum_name_list.index(want_colum_name)
        return self.find_colum(keyword_colum_name, keyword)[idx]

    def find_colum(self, colum_name, keyword):
        col = self.int_to_alpha[self.colum_name_list.index(colum_name) + 1]
        for i in range(3, 3 + self.row):
            if str(self.ws[col + str(i)].value) == keyword:
                idx = i
                break
        try:
            return self.table[idx - 3]
        except:
            return []

    def create_table(self):
        self.colum_name_list = self.serching_colum_name()
        if not self.colum_name_list:
            return
        row_downer = 3
        while True:
            if self.ws["C" + str(row_downer)].value == None:
                break
            row_downer += 1
        self.row = row_downer - 3
        for i in range(2, 2 + self.row):
            tmp = []
            for j in range(len(self.colum_name_list)):
                try:
                    tmp.append(self.normalize_string(self.ws[self.int_to_alpha[j + 1] + str(i + 1)].value))
                except:
                    tmp.append(self.ws[self.int_to_alpha[j + 1] + str(i + 1)].value)
            self.table.append(tmp)
        #print(self.table)

    def serching_colum_name(self):
        if self.ws['A2'].value != "注文方法":
            return []
        rng = self.ws['A2':'U2']
        colum_name = []
        for row in rng:
            for col in row:
                colum_name.append(col.value)
        return colum_name

    def normalize_string(self, s):
        return unicodedata.normalize('NFKC', s.strip())

    def find_colum_index(self, No):
        for i in range(3, 3 + self.row):
            if self.normalize_string(str(self.ws["C" + str(i)].value)) == self.normalize_string(No):
                return i
        return -1

    def items_count(self, idx):
        cnt = 0
        for i in range(7, 18 + 1):
            #print(self.ws.cell(idx, i).value, end=" ")
            try:
                cnt += int(self.normalize_string(str(self.ws.cell(idx, i).value)))
            except:
                pass
        return cnt

    def extra_row(self, No):
        idx = self.find_colum_index(No)
        tmp = []
        for i in range(1, 21 + 1):
            #print(self.ws.cell(idx, i).value, end = " ")
            #時刻の正規化
            if i == 6:
                tmp.append(self.ws.cell(idx, i).value[0:5])    
            else:
                tmp.append(self.ws.cell(idx, i).value)
        return tmp

    def change(self, No, after):
        idx = self.find_colum_index(No)
        if idx == -1:
            return
        for i in range(len(after)):
            self.ws.cell(idx, i+1, after[i])
        self.ws.cell(idx, 6).number_format = 'HH:MM'
        try:
            self.wb.save(self.excel_name)
        except:
            return []
        subprocess.Popen(['start',self.excel_name], shell=True)
        return [1]


    def check(self, No):
        idx = self.find_colum_index(No)
        if idx == -1:
            return
        self.ws.cell(idx, 21, self.items_count(idx))
        try:
            self.wb.save(self.excel_name)
        except:
            sg.print("Excelを閉じてから行ってください")
        subprocess.Popen(['start',self.excel_name], shell=True)

    def cancel(self, No):
        idx = self.find_colum_index(No)
        if idx == -1:
            return
        for i in range(idx, 3 + self.row):
            for j in range(1, len(self.colum_name_list)):
                new_cell = self.ws[self.int_to_alpha[j] + str(i + 1)].value
                if new_cell:
                    self.ws.cell(i, j, self.ws[self.int_to_alpha[j] + str(i + 1)].value)
                else:
                    self.ws.cell(i, j, "")
        self.row -= 1
        try:
            self.wb.save(self.excel_name)
        except:
            sg.print("Excelを閉じてから行ってください")
        subprocess.Popen(['start',self.excel_name], shell=True)


class GUIoutput():
    def __init__(self, column):
        self.column = column

    def start(self):
        layout = [
            [sg.Text("No.を教えてください"), sg.InputText(key = "No")],
            [sg.Button("検索", size=(5,2)), sg.Button("Excelを開く")]
        ]
        window = sg.Window("商品検索", layout, keep_on_top=True)
        while True:
            event, values = window.read()
            if event == sg.WINDOW_CLOSED:
                break
            if event == "検索":
                self.output_first(values)
            if event == "Excelを開く":
                self.open_excel()
                
    def open_excel(self):
        sercher = Sercher()
        sercher.open()

    def output_first(self, values):
        sercher = Sercher()
        sercher.create_table()
        if not sercher.table:
            sg.print("本日予約のExcelを開いて保存してください。")
            return
        if not values["No"]:
            sg.print("No.を入力してください")
            return
        items = sercher.data_mining("No.", values["No"])
        if not items:
            sg.print("このNo.は見つかりませんでした。")
            return
        No = values["No"]
        self.output_second(items, sercher, No)

    def output_second(self, dir, sercher, No):
        No_frame = sg.Frame("No.",[
            [sg.Text(dir["No."])]
        ])
        name_frame = sg.Frame("氏名",[
            [sg.Text(dir["名前"])]
        ])
        tmp = []
        for i in dir["cake"]:
            tmp.append([sg.Text(i + ' × ' + str(dir["cake"][i]))])
        if not tmp:
            items_frame = sg.Frame("商品", [
                [sg.Text("No." + str(dir["No."]) + " の商品はありません")]
            ])
        else:
            items_frame = sg.Frame("商品", tmp)
        ex_frame = sg.Frame("備考",[
            [sg.Text(dir["備考"])]
        ])
        frame_second = sg.Column([
            [items_frame],
            [ex_frame]
        ])
        layout = [
            [No_frame, name_frame],
            [frame_second],
            [sg.Button("キャンセル"), sg.Button("変更"), sg.Button("お渡し完了")],
            [sg.Button("閉じる")]
        ]
        window = sg.Window("お客様情報", layout, keep_on_top=True)
        while True:
            event, values = window.read()
            if event == sg.WINDOW_CLOSED:
                window.close()
                break
            if event == "キャンセル":
                sg.print("後日に予約を変更する場合は後日のExcelシートに新たに加えてください")
                sercher.cancel(No)
            if event == "お渡し完了":
                sercher.check(No)
            if event == "変更":
                self.output_change(No, sercher)
            if event == "閉じる":
                window.close()
                break

    def output_change(self, No, sercher):
        layout = []
        after = sercher.extra_row(No)
        for i in range(len(self.column)):
            layout.append(
                [
                    sg.Text(self.column[i]), 
                    sg.InputText(default_text=after[i], key= str(i) + 'it')
                ]
            )
        layout.append([sg.Button("OK")])
        window = sg.Window("変更", layout, keep_on_top=True)
        while True:
            event, values = window.read()
            #print(values)
            if event == sg.WINDOW_CLOSED:
                window.close()
                break
            if event == "OK":
                tmp = []
                for i in values:
                    tmp.append(values[i])
                tmp.pop(0)
                for i in tmp:
                    try:
                        hoge = datetime.strptime(i, "%H:%M").time()
                    except:
                        pass
                for i in range(len(tmp)):
                    try:
                        tmp[i] = int(tmp[i])
                    except:
                        pass
                bl = sercher.change(No, tmp)
                if bl:
                    window.close()
                    break
                else:
                    sg.print("Excelを閉じてから行ってください")

                
        
if __name__ == "__main__":
    #商品のみ１２品以内なら好きに設定できます。下記を変更してください。余った枠は""と空にしておいてください。
    column = ["注文方法", "会計未済※", "No.", "名前", "電話番号", "受取時間", "short4", "short5", "short6", "short7", "choco4", "choco5", "choco6", "choco7", "スノーモンブラン", "リッチショコラ", "紅タル", "フロマージュ", "計", "備考", "check"]
    guioutput = GUIoutput(column)
    guioutput.start()
